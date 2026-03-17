from __future__ import annotations

from pathlib import Path

import numpy as np
from scipy import stats

SHORT_NAMES: dict[str, str] = {
    "uci_aids_clinical_trials_group_study_175": "AIDS",
    "uci_adult": "Adult",
    "uci_bank_marketing": "Bank",
    "uci_banknote_authentication": "Banknote",
    "uci_census_income": "Census",
    "uci_contraceptive_method_choice": "CMC",
    "uci_habermans_survival": "Haberman",
    "uci_heart_disease": "Heart",
    "uci_occupancy_detection": "Occupancy",
    "uci_risk_factor_prediction_of_chronic_kidney_disease": "Kidney",
    "uci_spectf_heart": "SPECTF",
    "uci_statlog_(german_credit_data)": "German",
}

METHOD_DISPLAY_NAMES: dict[str, str] = {
    "gemini": "Gemini",
    "gpt": "GPT",
    "claude": "Claude",
}


def _safe_nanmean(arr: np.ndarray) -> float:
    vals = np.asarray(arr, dtype=float)
    vals = vals[~np.isnan(vals)]
    if vals.size == 0:
        return float("nan")
    return float(np.mean(vals))


def _safe_nanstd(arr: np.ndarray) -> float:
    vals = np.asarray(arr, dtype=float)
    vals = vals[~np.isnan(vals)]
    if vals.size <= 1:
        return float("nan")
    return float(np.std(vals, ddof=1))


def _format_mean_std(arr: np.ndarray) -> str:
    mean_val = _safe_nanmean(arr)
    std_val = _safe_nanstd(arr)
    if np.isnan(mean_val):
        return "NaN"
    if np.isnan(std_val):
        return f"{mean_val:.3f} ± NaN"
    return f"{mean_val:.3f} ± {std_val:.3f}"


def _compute_pvalue(best_accs: np.ndarray, other_accs: np.ndarray) -> float:
    diff = np.asarray(best_accs, dtype=float) - np.asarray(other_accs, dtype=float)
    valid = ~np.isnan(diff) & (diff != 0.0)
    if int(np.sum(valid)) < 10:
        return 1.0
    try:
        _stat, pval = stats.wilcoxon(diff[valid], alternative="two-sided")
    except Exception:
        pval = 1.0
    return float(pval)


def write_ablation_excel(
    *,
    results: dict[str, dict[str, dict[int, dict[str, np.ndarray]]]],
    kernels: list[str],
    methods: list[str],
    v_ntras: list[int],
    datasets: list[str],
    output_path: Path,
) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    hdr_font = Font(bold=True)
    center = Alignment(horizontal="center")

    for ntras in v_ntras:
        ws_acc = wb.create_sheet(f"Accuracy_n{ntras}")
        ws_p = wb.create_sheet(f"Pvalue_n{ntras}")

        method_headers = [METHOD_DISPLAY_NAMES.get(method, method) for method in methods]
        headers = ["Kernel", "Dataset", *method_headers]
        for col_idx, header in enumerate(headers, start=1):
            c1 = ws_acc.cell(row=1, column=col_idx, value=header)
            c1.font = hdr_font
            c1.alignment = center
            c2 = ws_p.cell(row=1, column=col_idx, value=header)
            c2.font = hdr_font
            c2.alignment = center

        row = 2
        for kernel in kernels:
            for dataset in datasets:
                short_name = SHORT_NAMES.get(dataset, dataset.removeprefix("uci_"))
                ws_acc.cell(row=row, column=1, value=kernel)
                ws_acc.cell(row=row, column=2, value=short_name)
                ws_p.cell(row=row, column=1, value=kernel)
                ws_p.cell(row=row, column=2, value=short_name)

                mean_values = [
                    _safe_nanmean(results[kernel][method][ntras][dataset]) for method in methods
                ]
                best_idx = int(np.nanargmax(np.asarray(mean_values, dtype=float)))
                best_method = methods[best_idx]
                best_accs = results[kernel][best_method][ntras][dataset]

                for j, method in enumerate(methods, start=3):
                    arr = results[kernel][method][ntras][dataset]
                    ws_acc.cell(row=row, column=j, value=_format_mean_std(arr))

                    if method == best_method:
                        pvalue_cell = ws_p.cell(row=row, column=j, value="-")
                        pvalue_cell.font = Font(bold=True)
                    else:
                        pval = _compute_pvalue(best_accs, arr)
                        cell_value = "-" if pval > 0.05 else f"{pval:.6f}"
                        ws_p.cell(row=row, column=j, value=cell_value)

                row += 1

        for ws in [ws_acc, ws_p]:
            ws.column_dimensions["A"].width = 10
            ws.column_dimensions["B"].width = 16
            for ci in range(3, 3 + len(methods)):
                ws.column_dimensions[get_column_letter(ci)].width = 18

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
