import os
import sys
from datetime import datetime
from pathlib import Path

import numpy as np
from openpyxl import Workbook

current_dir = os.path.dirname(os.path.abspath(__file__))
src_path = os.path.join(current_dir, "../src")
sys.path.append(src_path)

import kernels
import pfgd_train
import preprocessing


def run_experiment():
    np.random.seed(0)

    v_nfeas = np.unique(np.floor(np.logspace(2, 4, 9))).astype(int)

    qpmeth1s = ["cvxopt", "proxqp", "cqkp"]
    n_meths = len(qpmeth1s)
    n_repeats = 3
    avg_points = (2, 3)
    tmmat1 = np.full((len(v_nfeas), n_meths), np.nan)
    tmmat_repeats = np.full((n_repeats, len(v_nfeas), n_meths), np.nan)
    avg_mats = {
        k: np.full((len(v_nfeas), n_meths), np.nan, dtype=float) for k in avg_points
    }
    timestamp = datetime.now().strftime("%y%m%d%H%M%S")
    output_dir = Path("result")
    output_dir.mkdir(parents=True, exist_ok=True)

    base_name = f"{timestamp}_runtime"
    csv_path = output_dir / f"{base_name}.csv"
    xlsx_path = output_dir / f"{base_name}.xlsx"
    npz_path = output_dir / f"{base_name}.npz"
    header_cols = ["nfeas", *qpmeth1s]
    current_nfeas = None
    current_qpmeth = None

    def write_runtime_csv_snapshot():
        with csv_path.open("w", encoding="utf-8") as f:
            f.write(",".join(header_cols) + "\n")
            for i, nfeas in enumerate(v_nfeas):
                vals = [f"{tmmat1[i, j]:.10f}" for j in range(n_meths)]
                f.write(f"{nfeas}," + ",".join(vals) + "\n")

    def append_runtime_csv_stop_reason(reason_text):
        reason_text = str(reason_text).replace("\n", " ").replace("\r", " ")
        with csv_path.open("a", encoding="utf-8") as f:
            f.write("\n")
            f.write(f"# stopped_at,{datetime.now().isoformat(timespec='seconds')}\n")
            f.write(f"# stop_reason,{reason_text}\n")

    def update_avg_mats_for_cell(i_nfeas, i_qpmeth):
        vals = tmmat_repeats[:, i_nfeas, i_qpmeth]
        for k in avg_points:
            first_k = vals[:k]
            if np.all(np.isfinite(first_k)):
                avg_mats[k][i_nfeas, i_qpmeth] = float(np.mean(first_k))

    def update_current_mean_for_cell(i_nfeas, i_qpmeth, rep_count):
        vals = tmmat_repeats[:rep_count, i_nfeas, i_qpmeth]
        vals = vals[np.isfinite(vals)]
        if vals.size > 0:
            tmmat1[i_nfeas, i_qpmeth] = float(np.mean(vals))

    def print_average_table(mat, label):
        print(f"\n=== {label} (s) ===")
        header = f"{'nfeas':>10s}"
        for m in qpmeth1s:
            header += f"  {m + ' (s)':>12s}"
        print(header)
        print("-" * (10 + 14 * n_meths))
        for j in range(len(v_nfeas)):
            row = f"{v_nfeas[j]:10d}"
            for k in range(n_meths):
                row += f"  {mat[j, k]:12.4f}"
            print(row)

    def write_runtime_xlsx_snapshot():
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        ws_first = wb.create_sheet(title="avg_1")
        ws_first.append(["nfeas", *qpmeth1s])
        mat_first = tmmat_repeats[0, :, :]
        for i, nfeas in enumerate(v_nfeas):
            row = [int(nfeas)]
            for j in range(n_meths):
                val = mat_first[i, j]
                row.append(None if not np.isfinite(val) else float(val))
            ws_first.append(row)
        for i in range(2, len(v_nfeas) + 2):
            for j in range(2, n_meths + 2):
                cell = ws_first.cell(row=i, column=j)
                if cell.value is not None:
                    cell.number_format = "0.0000"

        for k in avg_points:
            ws = wb.create_sheet(title=f"avg_{k}")
            ws.append(["nfeas", *qpmeth1s])
            mat = avg_mats[k]
            for i, nfeas in enumerate(v_nfeas):
                row = [int(nfeas)]
                for j in range(n_meths):
                    val = mat[i, j]
                    row.append(None if not np.isfinite(val) else float(val))
                ws.append(row)
            for i in range(2, len(v_nfeas) + 2):
                for j in range(2, n_meths + 2):
                    cell = ws.cell(row=i, column=j)
                    if cell.value is not None:
                        cell.number_format = "0.0000"

        ws_raw = wb.create_sheet(title="raw_repeat")
        ws_raw.append(["repeat", "nfeas", *qpmeth1s])
        for r in range(n_repeats):
            for i, nfeas in enumerate(v_nfeas):
                row = [r + 1, int(nfeas)]
                for j in range(n_meths):
                    val = tmmat_repeats[r, i, j]
                    row.append(None if not np.isfinite(val) else float(val))
                ws_raw.append(row)
        for i in range(2, 2 + n_repeats * len(v_nfeas)):
            for j in range(3, n_meths + 3):
                cell = ws_raw.cell(row=i, column=j)
                if cell.value is not None:
                    cell.number_format = "0.0000"

        wb.save(xlsx_path)

    def save_runtime_npz():
        payload = {
            "timestamp": timestamp,
            "v_nfeas": v_nfeas,
            "qpmeth1s": np.array(qpmeth1s, dtype=object),
            "tmmat1": tmmat1,
            "tmmat_repeats": tmmat_repeats,
            "tmmat_avg1": tmmat_repeats[0, :, :].copy(),
        }
        for k in avg_points:
            payload[f"tmmat_avg{k}"] = avg_mats[k]
        np.savez(npz_path, **payload)

    write_runtime_csv_snapshot()
    write_runtime_xlsx_snapshot()

    max_nfeas_by_method = {
        "proxqp": 10**4,
        "cvxopt": 10**4,
        "cqkp": np.inf,
    }
    run_plan = np.array(
        [
            [nfeas <= max_nfeas_by_method.get(meth, np.inf) for meth in qpmeth1s]
            for nfeas in v_nfeas
        ],
        dtype=bool,
    )
    n_total = int(run_plan.sum()) * n_repeats
    done = 0

    try:
        ntras = 100
        ntsts = 2
        lamn = 1.0
        gam_rbf = 0.001
        b_primes = np.array([1.0, -1.0])
        s_prime = np.array([1.0, 0.0])
        gam_sm = 4.0
        thres_gap = 1e-4

        for i_rep in range(n_repeats):
            for i_nfeas, nfeas in enumerate(v_nfeas):
                current_nfeas = int(nfeas)
                current_qpmeth = "preprocessing"
                npts = ntras + ntsts
                y_all = np.concatenate([np.ones(npts // 2), -np.ones(npts // 2)])

                X_all = np.outer(np.ones(nfeas), y_all) * 0.1
                col_norms = np.sqrt(np.sum(X_all**2, axis=0))
                X_all = X_all * (1.0 / col_norms)

                cvec_sgn = np.ones(nfeas)
                l_tra = np.concatenate(
                    [np.ones(ntras, dtype=bool), np.zeros(ntsts, dtype=bool)]
                )
                l_tst = ~l_tra
                X_tra = X_all[:, l_tra]
                X_tst = X_all[:, l_tst]
                y_tra = y_all[l_tra]

                lam1 = lamn / ntras

                def fh_kern(X1, X2, _gam=gam_rbf):
                    return kernels.rbf_kernel(X1, X2, _gam)

                K_xx, K_qx, c1, c2, sigvec1, K_qx_ekm, K_xx_ekm = (
                    preprocessing.prepro_sckmtwo_blocks_light(
                        X_tra,
                        y_tra,
                        X_tst,
                        fh_kern,
                        cvec_sgn,
                        s1=b_primes,
                        s2=s_prime,
                    )
                )
                kern_blocks = {"K_xx": K_xx, "K_qx": K_qx, "c1": c1, "c2": c2}
                for i_qpmeth, qpmeth1 in enumerate(qpmeth1s):
                    current_qpmeth = qpmeth1
                    if nfeas > max_nfeas_by_method.get(qpmeth1, np.inf):
                        continue

                    done += 1
                    bar_len = 30
                    filled = int(bar_len * done / max(n_total, 1))
                    bar = "=" * filled + "-" * (bar_len - filled)
                    print(
                        f"\r[{bar}] {done}/{n_total}  pass={i_rep + 1}/{n_repeats}, nfeas={nfeas}, {qpmeth1}...",
                        end="",
                        flush=True,
                    )
                    res1_pfgd = pfgd_train.train_pfgd(
                        kern_blocks,
                        sigvec1,
                        lam1,
                        gam_sm,
                        qpmeth1,
                        nepochs=100000,
                        thres_gap=thres_gap,
                        verbose=0,
                    )
                    tms_pfgd = res1_pfgd["tms"]
                    tm_val = float(tms_pfgd[-1]) if len(tms_pfgd) > 0 else 0.0
                    tmmat_repeats[i_rep, i_nfeas, i_qpmeth] = tm_val

                    update_avg_mats_for_cell(i_nfeas, i_qpmeth)
                    update_current_mean_for_cell(i_nfeas, i_qpmeth, i_rep + 1)
                    write_runtime_csv_snapshot()
                    write_runtime_xlsx_snapshot()

            rep_num = i_rep + 1
            if rep_num in avg_points:
                tmmat1[:, :] = avg_mats[rep_num]
                print("\r" + " " * 100 + "\r", end="")
                print_average_table(avg_mats[rep_num], f"{rep_num}回平均")
                write_runtime_csv_snapshot()
                write_runtime_xlsx_snapshot()
    except BaseException as exc:
        write_runtime_csv_snapshot()
        write_runtime_xlsx_snapshot()
        reason = f"{type(exc).__name__}: {exc}"
        if current_nfeas is not None:
            reason += f" (nfeas={current_nfeas}"
            if current_qpmeth is not None:
                reason += f", method={current_qpmeth}"
            reason += ")"
        append_runtime_csv_stop_reason(reason)
        save_runtime_npz()
        print("\n処理が中断されました。原因をCSV末尾に記録しました。")
        print(f"Saved CSV: {csv_path}")
        print(f"Saved XLSX: {xlsx_path}")
        print(f"Saved NPZ: {npz_path}")
        return 1
    final_k = max(avg_points)
    print("\r" + " " * 80 + "\r", end="")
    print_average_table(avg_mats[final_k], f"最終 {final_k}回平均")
    tmmat1 = avg_mats[final_k].copy()
    write_runtime_csv_snapshot()
    write_runtime_xlsx_snapshot()

    save_runtime_npz()

    print(f"Saved CSV: {csv_path}")
    print(f"Saved XLSX: {xlsx_path}")
    print(f"Saved NPZ: {npz_path}")

    return 0


if __name__ == "__main__":
    sys.exit(run_experiment())
