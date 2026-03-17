

from __future__ import annotations

import sys
import time
import warnings
from datetime import datetime
from pathlib import Path
import shutil

import numpy as np
from scipy import stats

REPO_ROOT = Path(__file__).resolve().parent.parent
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from src.demo1321_exp import (  # noqa: E402
    load_dataset_for_demo1321,
    prepare_demo1321_dataset,
)
from src.gbdt_baselines import (  # noqa: E402
    train_eval_lightgbm
)
from src.kernels import get_kernel_function  # noqa: E402
from src.pfgd_train import train_pfgd  # noqa: E402
from src.preprocessing import prepro_sckmtwo_blocks_light  # noqa: E402
from src.splitters import h34_gen_lmat_tra  # noqa: E402

# ── dataset short names for Excel headers ──────────────────────────
SHORT_NAMES: dict[str, str] = {
    "uci_aids_clinical_trials_group_study_175": "AIDS",
    "uci_adult": "Adult",
    "uci_bank_marketing": "Bank",
    "uci_banknote_authentication": "Banknote",
    "uci_census_income": "Census",
    "uci_contraceptive_method_choice": "CMC",
    "uci_habermans_survival": "Haberman",
    "uci_heart_disease": "Heart",
    "uci_occupancy_detection": "Occupancy",
    "uci_risk_factor_prediction_of_chronic_kidney_disease": "Kidney",
    "uci_spectf_heart": "SPECTF",
    "uci_statlog_(german_credit_data)": "German",
}


# ====================================================================
# helpers
# ====================================================================


def _configure_warning_filters() -> None:
    """Suppress known third-party warnings that do not affect results."""
    warnings.filterwarnings(
        "ignore",
        message="X does not have valid feature names, but LGBMClassifier was fitted with feature names",
        category=UserWarning,
    )


def _percv_acc(sco: np.ndarray, y: np.ndarray) -> float:
    """Classification accuracy for a *single* CV fold."""
    sco = sco.reshape(-1)
    y = y.reshape(-1)
    n_valid = int(np.sum(y != 0))
    if n_valid == 0:
        return float("nan")
    return float(np.sum(sco * y > 0) / n_valid)


# ====================================================================
# per-CV runners
# ====================================================================


def run_kernel_percv(
    X_all: np.ndarray,
    y_all: np.ndarray,
    cvec_sgn: np.ndarray,
    lmat_tra: np.ndarray,
    mode_sgncon: str,
    typ_kern: str,
    *,
    eta_nlr: float,
    gam_rbf: float,
    lamn: float,
    b_prime: np.ndarray,
    s_prime: np.ndarray,
    gam_sm: float,
    nepochs: int,
    qpmeth1: str,
) -> np.ndarray:
    """Run kernel PFGD for every CV fold, return per-CV acc array."""
    ncvs_rnd = lmat_tra.shape[0]
    fh_kern = get_kernel_function(typ_kern, eta_nlr=eta_nlr, gam_rbf=gam_rbf)
    accs = np.empty(ncvs_rnd, dtype=float)

    for cv in range(ncvs_rnd):
        l_tra = lmat_tra[cv]
        l_tst = ~l_tra
        X_tra, X_tst = X_all[:, l_tra], X_all[:, l_tst]
        y_tra, y_tst = y_all[l_tra], y_all[l_tst]

        ntras_cv = y_tra.size
        lam1 = lamn / ntras_cv

        cvec_for_prepro = np.zeros_like(cvec_sgn) if mode_sgncon == "sf" else cvec_sgn

        K_xx, K_qx, c1, c2, sigvec1, K_qx_ekm, K_xx_ekm = prepro_sckmtwo_blocks_light(
            X_tra,
            y_tra,
            X_tst,
            fh_kern,
            cvec_for_prepro,
            s1=b_prime,
            s2=s_prime,
        )

        res = train_pfgd(
            K_xx=K_xx,
            K_qx=K_qx,
            c1=c1,
            c2=c2,
            sigvec1=sigvec1,
            lam1=lam1,
            gam_sm=gam_sm,
            qpmeth1=qpmeth1,
            nepochs=nepochs,
            mode_sgncon=mode_sgncon,
            verbose=0,
        )

        alph1 = res["alph1"]
        sigbeta1 = res["sigbeta1"]
        ilamn1 = 1.0 / (lam1 * ntras_cv)
        yalph1 = y_tra * alph1
        sco_tst = ilamn1 * (K_xx_ekm.T @ yalph1)
        if mode_sgncon == "sc":
            sco_tst = sco_tst + K_qx_ekm.T @ sigbeta1

        accs[cv] = _percv_acc(sco_tst, y_tst)

    return accs


def run_gbdt_percv(
    X_all: np.ndarray,
    y_all: np.ndarray,
    cvec_sgn: np.ndarray,
    lmat_tra: np.ndarray,
    mode_sgncon: str,
    gbdt_type: str,
) -> np.ndarray:
    """Run LightGBM for every CV fold, return per-CV acc."""
    ncvs_rnd = lmat_tra.shape[0]
    accs = np.empty(ncvs_rnd, dtype=float)

    eval_fn = train_eval_lightgbm

    for cv in range(ncvs_rnd):
        l_tra = lmat_tra[cv]
        l_tst = ~l_tra
        X_tra, X_tst = X_all[:, l_tra], X_all[:, l_tst]
        y_tra, y_tst = y_all[l_tra], y_all[l_tst]

        try:
            sco_tst = eval_fn(X_tra, y_tra, X_tst, cvec_sgn, mode_sgncon)
            accs[cv] = _percv_acc(sco_tst, y_tst)
        except Exception as exc:  # noqa: BLE001
            print(f"      [WARN] cv={cv} failed ({exc!r}), setting acc=NaN")
            accs[cv] = float("nan")

    return accs


# ====================================================================
# Excel writer
# ====================================================================


def write_excel(
    results: dict[str, dict[int, dict[str, np.ndarray]]],
    all_meths: list[str],
    v_ntras: list[int],
    dbname1s: list[str],
    output_dir: Path,
) -> Path:
    """Build the formatted Excel workbook and return its path."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    n_meths = len(all_meths)
    n_n = len(v_ntras)
    n_db = len(dbname1s)

    # ── collect statistics ──────────────────────────────────────────
    means = np.full((n_meths, n_n, n_db), np.nan)
    stds = np.full((n_meths, n_n, n_db), np.nan)
    percv: dict[tuple[int, int, int], np.ndarray] = {}

    for i_m, meth in enumerate(all_meths):
        for i_n, ntras in enumerate(v_ntras):
            for i_db, db in enumerate(dbname1s):
                arr = results[meth][ntras][db]
                means[i_m, i_n, i_db] = np.nanmean(arr)
                stds[i_m, i_n, i_db] = np.nanstd(arr, ddof=1)
                percv[(i_m, i_n, i_db)] = arr

    # ── statistical tests (column-wise: best vs each other) ────────
    is_best = np.zeros((n_meths, n_n, n_db), dtype=bool)
    p_wilcoxon = np.ones((n_meths, n_n, n_db))
    p_ttest = np.ones((n_meths, n_n, n_db))
    not_sig_diff = np.zeros((n_meths, n_n, n_db), dtype=bool)

    for i_n in range(n_n):
        for i_db in range(n_db):
            col_means = means[:, i_n, i_db]
            best_idx = int(np.nanargmax(col_means))
            is_best[best_idx, i_n, i_db] = True
            best_accs = percv[(best_idx, i_n, i_db)]

            for i_m in range(n_meths):
                if i_m == best_idx:
                    not_sig_diff[i_m, i_n, i_db] = True
                    continue

                other_accs = percv[(i_m, i_n, i_db)]
                diff = best_accs - other_accs
                valid = ~np.isnan(diff) & (diff != 0.0)

                if int(np.sum(valid)) < 10:
                    # Too few non-zero differences → treat as tied
                    not_sig_diff[i_m, i_n, i_db] = True
                    continue

                # Wilcoxon signed-rank test
                try:
                    _, pw = stats.wilcoxon(diff[valid], alternative="two-sided")
                    p_wilcoxon[i_m, i_n, i_db] = pw
                except Exception:
                    pw = 1.0
                    p_wilcoxon[i_m, i_n, i_db] = pw

                # Paired t-test (one-sample on diff)
                try:
                    _, pt = stats.ttest_1samp(diff[~np.isnan(diff)], 0.0)
                    p_ttest[i_m, i_n, i_db] = pt
                except Exception:
                    pt = 1.0
                    p_ttest[i_m, i_n, i_db] = pt

                not_sig_diff[i_m, i_n, i_db] = pw > 0.05

    # ── create workbook ─────────────────────────────────────────────
    wb = Workbook()

    # ---- Sheet 1: Results ----
    ws = wb.active
    ws.title = "Results"

    hdr_font = Font(bold=True)
    hdr_align = Alignment(horizontal="center")
    head_align = Alignment(horizontal="center")

    # Row 1: n values (merged)
    ws.cell(row=1, column=1, value="Method").font = hdr_font
    col = 2
    for ntras in v_ntras:
        ws.merge_cells(
            start_row=1,
            start_column=col,
            end_row=1,
            end_column=col + n_db - 1,
        )
        c = ws.cell(row=1, column=col, value=f"n = {ntras}")
        c.font = hdr_font
        c.alignment = hdr_align
        col += n_db

    # Row 2: dataset short names
    col = 2
    for _ntras in v_ntras:
        for db in dbname1s:
            short = SHORT_NAMES.get(db, db.replace("uci_", ""))
            c = ws.cell(row=2, column=col, value=short)
            c.font = hdr_font
            c.alignment = head_align
            col += 1

    # Data rows
    for i_m, meth in enumerate(all_meths):
        row = 3 + i_m
        ws.cell(row=row, column=1, value=meth)
        col = 2
        for i_n in range(n_n):
            for i_db in range(n_db):
                m = means[i_m, i_n, i_db]
                s = stds[i_m, i_n, i_db]
                cell = ws.cell(row=row, column=col, value=f"{m:.3f} ± {s:.3f}")

                bold = bool(is_best[i_m, i_n, i_db])
                underline = bool(not_sig_diff[i_m, i_n, i_db]) and not bold

                if bold and underline:
                    cell.font = Font(bold=True, underline="single")
                elif bold:
                    cell.font = Font(bold=True)
                elif underline:
                    cell.font = Font(underline="single")

                col += 1

    # Column widths
    ws.column_dimensions["A"].width = 18
    for ci in range(2, 2 + n_n * n_db):
        ws.column_dimensions[get_column_letter(ci)].width = 16

    # ---- Sheet 2: p-values (Wilcoxon) ----
    _write_pval_sheet(
        wb, "p-values (Wilcoxon)", p_wilcoxon, all_meths, v_ntras, dbname1s
    )

    # ---- Sheet 3: p-values (t-test) ----
    _write_pval_sheet(wb, "p-values (t-test)", p_ttest, all_meths, v_ntras, dbname1s)

    # ---- save ----
    timestamp = datetime.now().strftime("%y%m%d%H%M%S")
    out_path = output_dir / f"results_{timestamp}.xlsx"
    wb.save(out_path)
    return out_path


def _write_pval_sheet(
    wb,
    title: str,
    pvals: np.ndarray,
    all_meths: list[str],
    v_ntras: list[int],
    dbname1s: list[str],
) -> None:
    """Helper: write a p-value matrix to a new worksheet."""
    from openpyxl.styles import Alignment, Font

    ws = wb.create_sheet(title)
    hdr_font = Font(bold=True)
    n_n = len(v_ntras)
    n_db = len(dbname1s)

    # Header
    ws.cell(row=1, column=1, value="Method").font = hdr_font
    col = 2
    for ntras in v_ntras:
        for db in dbname1s:
            short = SHORT_NAMES.get(db, db.replace("uci_", ""))
            c = ws.cell(row=1, column=col, value=f"n={ntras} {short}")
            c.font = hdr_font
            c.alignment = Alignment(horizontal="center")
            col += 1

    # Data
    for i_m, meth in enumerate(all_meths):
        row = 2 + i_m
        ws.cell(row=row, column=1, value=meth)
        col = 2
        for i_n in range(n_n):
            for i_db in range(n_db):
                ws.cell(
                    row=row, column=col, value=round(float(pvals[i_m, i_n, i_db]), 6)
                )
                col += 1


# ====================================================================
# main
# ====================================================================


def main() -> None:
    _configure_warning_filters()
    np.random.seed(0)

    # ── experiment configuration (matches demo1321_01.py) ──────────
    ncvs_rnd = 200
    rng_init = 0
    data_dir = REPO_ROOT / "data"

    dbname1s: list[str] = sorted(p.stem for p in data_dir.glob("*.mat"))
    if not dbname1s:
        raise FileNotFoundError(f"No dataset .mat files found under: {data_dir}")

    v_ntras: list[int] = [5, 10]

    # kernel hyper-parameters (same as demo1321_01)
    scale = 10.0
    bias_factor = 10.0
    lamn = 1.0
    delta = 1.0
    eta_nlr = 0.8
    gam_rbf = 0.001
    b_prime = np.array([1.0, -1.0])
    s_prime = np.array([delta, 0.0])
    gam_sm = 0.01
    nepochs = 1000
    qpmeth1 = "cqkp"

    # methods ---------------------------------------------------------
    kernel_meths = [
        "sc-lin",
        "sc-rbf",
        "sf-lin",
        "sf-rbf",
    ]
    gbdt_meths = [
        "mc-lightgbm",
        "mf-lightgbm",
    ]
    all_meths = kernel_meths + gbdt_meths

    # ── results container: meth → ntras → dbname → ndarray(200,) ───
    results: dict[str, dict[int, dict[str, np.ndarray]]] = {
        m: {n: {} for n in v_ntras} for m in all_meths
    }

    # ── output directory ────────────────────────────────────────────
    timestamp = datetime.now().strftime("%y%m%d%H%M%S")
    output_dir = REPO_ROOT / "out-demo1321-02" / timestamp
    output_dir.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(Path(__file__), output_dir / Path(__file__).name)
    print(f"Output directory: {output_dir}")

    total_t0 = time.time()

    # ================================================================
    #  main loop
    # ================================================================
    for i_db, dbname1 in enumerate(dbname1s, start=1):
        print(f"\n[{i_db:2d}/{len(dbname1s):2d}] {dbname1}")
        param_dataset = load_dataset_for_demo1321(str(data_dir), dbname1, verbose=False)

        prepared = prepare_demo1321_dataset(
            param_dataset=param_dataset,
            dbname1=dbname1,
            scale=scale,
            bias_factor=bias_factor,
        )
        X_all = prepared["X_all"]
        y_all = prepared["y_all"]
        cvec_sgn = prepared["cvec_sgn"]

        for ntras in v_ntras:
            # — generate CV splits (deterministic, same as demo1321_01) —
            np.random.seed(rng_init)
            lmat_tra = h34_gen_lmat_tra(y_all, ntras, ncvs_rnd)

            # ---------- kernel methods ----------
            common_kw = dict(
                eta_nlr=eta_nlr,
                gam_rbf=gam_rbf,
                lamn=lamn,
                b_prime=b_prime,
                s_prime=s_prime,
                gam_sm=gam_sm,
                nepochs=nepochs,
                qpmeth1=qpmeth1,
            )

            for meth in kernel_meths:
                mode_sgncon = meth[:2]
                typ_kern = meth[3:]
                t0 = time.time()
                accs = run_kernel_percv(
                    X_all,
                    y_all,
                    cvec_sgn,
                    lmat_tra,
                    mode_sgncon,
                    typ_kern,
                    **common_kw,
                )
                elapsed = time.time() - t0
                results[meth][ntras][dbname1] = accs
                print(
                    f"  n={ntras:3d}  {meth:16s}  "
                    f"acc={np.nanmean(accs):.4f}±{np.nanstd(accs):.4f}  "
                    f"({elapsed:.1f}s)"
                )

            # ---------- GBDT methods ----------
            for meth in gbdt_meths:
                mode_sgncon = meth[:2]
                gbdt_type = meth[3:]
                t0 = time.time()
                accs = run_gbdt_percv(
                    X_all,
                    y_all,
                    cvec_sgn,
                    lmat_tra,
                    mode_sgncon,
                    gbdt_type,
                )
                elapsed = time.time() - t0
                results[meth][ntras][dbname1] = accs
                print(
                    f"  n={ntras:3d}  {meth:16s}  "
                    f"acc={np.nanmean(accs):.4f}±{np.nanstd(accs):.4f}  "
                    f"({elapsed:.1f}s)"
                )

    total_elapsed = time.time() - total_t0
    print(f"\nAll experiments done in {total_elapsed:.0f}s.")

    # ── save raw per-CV accs ----------------------------------------
    npz_path = output_dir / "percv_accs.npz"
    flat: dict[str, np.ndarray] = {}
    for meth in all_meths:
        for ntras in v_ntras:
            for db in dbname1s:
                key = f"{meth}__n{ntras}__{db}"
                flat[key] = results[meth][ntras][db]
    np.savez_compressed(npz_path, **flat)
    print(f"Raw accs saved: {npz_path}")

    # ── write Excel --------------------------------------------------
    xlsx_path = write_excel(results, all_meths, v_ntras, dbname1s, output_dir)
    print(f"Excel saved:    {xlsx_path}")


if __name__ == "__main__":
    main()
